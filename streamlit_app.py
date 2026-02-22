# -*- coding: utf-8 -*-
"""
개인 자산관리 - Streamlit 웹 앱
─────────────────────────────────────────────────────
데이터 : GitHub private repo 에서 다운로드G
인증   : streamlit-authenticator (bcrypt 해시)
─────────────────────────────────────────────────────
"""

import io
import re
import base64
import tempfile
import warnings
from pathlib import Path
from html.parser import HTMLParser

warnings.filterwarnings('ignore')

import requests
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pandas as pd
import streamlit as st
import streamlit_authenticator as stauth
import yaml
from yaml.loader import SafeLoader

# ── 한글 폰트 (Linux: NanumGothic, Windows: Malgun Gothic) ─────
import matplotlib.font_manager as fm
for _fc in ['NanumGothic', 'Malgun Gothic', 'AppleGothic', 'DejaVu Sans']:
    if any(_fc.lower() in f.name.lower() for f in fm.fontManager.ttflist):
        plt.rcParams['font.family'] = _fc
        break
plt.rcParams['axes.unicode_minus'] = False

# ── 팔레트 ──────────────────────────────────────────────────────
PERSON_COLORS  = {'박수환': '#4472C4', '황선미': '#ED7D31'}
DEFAULT_COLORS = ['#4472C4', '#ED7D31', '#A9D18E', '#FFC000',
                  '#9E63A0', '#00B0F0', '#FF6347', '#5A5A5A']

# ════════════════════════════════════════════════════════════════
# 데이터 파싱 (asset_manager.py 와 동일 로직 - tkinter 의존 없음)
# ════════════════════════════════════════════════════════════════

def _to_float(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ('-', '', 'N/A', 'n/a'):
        return None
    is_pct = s.endswith('%')
    try:
        n = float(s.replace(',', '').replace('%', ''))
        return n / 100 if is_pct else n
    except ValueError:
        return None

def _n(val):
    if val is None:
        return None
    s = str(val).strip().replace(',', '').replace('+', '')
    if s in ('-', '', 'N/A', 'nan'):
        return None
    is_pct = s.endswith('%')
    try:
        v = float(s.replace('%', ''))
        return v / 100 if is_pct else v
    except ValueError:
        return None

def _dash(val):
    v = _n(val)
    return v if v is not None else '-'

CLEAN_COLS = ['상품명', '보유수량', '현재가', '평균매입가',
              '매입금액', '평가금액', '평가손익', '손익률']


class _HtmlTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables = []
        self._stack = []
        self._row   = []
        self._cell  = []
        self._in_cell = False

    def handle_starttag(self, tag, attrs):
        if tag == 'table':   self._stack.append([])
        elif tag == 'tr':    self._row = []
        elif tag in ('td','th'): self._cell = []; self._in_cell = True

    def handle_endtag(self, tag):
        if tag == 'table':
            if self._stack:
                tbl = self._stack.pop()
                if tbl: self.tables.append(tbl)
        elif tag == 'tr':
            if self._row and self._stack: self._stack[-1].append(self._row)
            self._row = []
        elif tag in ('td','th'):
            self._row.append(' '.join(self._cell).strip())
            self._in_cell = False

    def handle_data(self, data):
        if self._in_cell:
            s = data.strip()
            if s: self._cell.append(s)


def parse_mirae_xls(path: Path, account: str) -> pd.DataFrame:
    content = None
    for enc in ('utf-8', 'cp949', 'euc-kr'):
        try:
            content = path.read_text(encoding=enc, errors='strict'); break
        except (UnicodeDecodeError, LookupError):
            continue
    if content is None:
        content = path.read_text(encoding='cp949', errors='replace')
    p = _HtmlTableParser()
    p.feed(content)
    tables = p.tables
    target_idx = None
    for i, tbl in enumerate(tables):
        flat = ' '.join(c for row in tbl for c in row)
        if '상품보유현황' in flat and account in flat:
            target_idx = i + 1; break
    if target_idx is None or target_idx >= len(tables):
        for i, tbl in enumerate(tables):
            if tbl and tbl[0] and '상품명' in tbl[0]:
                target_idx = i; break
    if target_idx is None or target_idx >= len(tables):
        return pd.DataFrame(columns=CLEAN_COLS)
    data_tbl = tables[target_idx]
    if not data_tbl: return pd.DataFrame(columns=CLEAN_COLS)
    header = data_tbl[0]
    if '상품명' not in header: return pd.DataFrame(columns=CLEAN_COLS)
    col_idx = {h: j for j, h in enumerate(header)}
    def get(row, col):
        idx = col_idx.get(col)
        return row[idx] if idx is not None and idx < len(row) else None
    rows = []
    for row in data_tbl[1:]:
        nm = get(row, '상품명')
        if not nm or not nm.strip(): continue
        rows.append({'상품명': nm.strip(), '보유수량': _dash(get(row,'보유수량')),
                     '현재가': _dash(get(row,'현재가')), '평균매입가': _dash(get(row,'평균매입가')),
                     '매입금액': _dash(get(row,'매입금액')), '평가금액': _dash(get(row,'평가금액')),
                     '평가손익': _dash(get(row,'평가손익')), '손익률': _dash(get(row,'손익률'))})
    return pd.DataFrame(rows, columns=CLEAN_COLS)


def parse_samsung_xlsx(path: Path) -> pd.DataFrame:
    from openpyxl import load_workbook as _lw
    wb = _lw(path, data_only=True)
    ws = wb.active
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    cash = 0.0
    for row in all_rows:
        cell0 = str(row[0] or '').strip()
        if 'KRW' in cell0 and '한국' in cell0:
            v = _n(row[1]); cash = v if v else 0.0; break
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if len(row) > 2 and str(row[2] or '').strip() == '종목명':
            header_row_idx = i; break
    if header_row_idx is None: return pd.DataFrame(columns=CLEAN_COLS)
    result = []
    for row in all_rows[header_row_idx + 1:]:
        nm = str(row[2] or '').strip() if len(row) > 2 else ''
        if not nm: continue
        qty  = _n(row[3]) if len(row) > 3 else None
        cost = _n(row[4]) if len(row) > 4 else None
        val  = _n(row[5]) if len(row) > 5 else None
        pnl  = _n(row[6]) if len(row) > 6 else None
        pct_v= _n(row[9]) if len(row) > 9 else None
        if cost is None and val is None: continue
        avg = round(cost / qty) if (cost and qty) else None
        px  = round(val  / qty) if (val  and qty) else None
        pct = (pct_v / 100) if pct_v is not None else (
              (pnl / cost) if (pnl and cost) else None)
        result.append({'상품명': nm, '보유수량': qty if qty is not None else '-',
                       '현재가': px if px is not None else '-',
                       '평균매입가': avg if avg is not None else '-',
                       '매입금액': cost if cost is not None else '-',
                       '평가금액': val if val is not None else '-',
                       '평가손익': pnl if pnl is not None else '-',
                       '손익률': pct if pct is not None else '-'})
    if cash > 0:
        result.append({'상품명':'현금잔고','보유수량':'-','현재가':'-','평균매입가':'-',
                       '매입금액':cash,'평가금액':cash,'평가손익':'-','손익률':'-'})
    return pd.DataFrame(result, columns=CLEAN_COLS)


def parse_kiwoom_csv(path: Path) -> pd.DataFrame:
    df = None
    for enc in ('cp949', 'utf-8-sig', 'utf-8'):
        try: df = pd.read_csv(path, skiprows=1, encoding=enc, dtype=str); break
        except UnicodeDecodeError: continue
    if df is None: return pd.DataFrame(columns=CLEAN_COLS)
    df.columns = [str(c).strip() for c in df.columns]
    result = []
    for _, r in df.iterrows():
        nm = str(r.get('종목명','')).strip()
        if not nm or nm.lower() == 'nan': continue
        qty=_n(r.get('보유량')); px_usd=_n(r.get('현재가')); avg_usd=_n(r.get('매입가'))
        cost_usd=_n(r.get('매입금액')); val_krw=_n(r.get('환전평가금액'))
        buy_rate=_n(r.get('매입환율')); cur_rate=_n(r.get('현재환율'))
        if None in (qty, cost_usd, buy_rate, cur_rate): continue
        cost_krw=round(cost_usd*buy_rate); px_krw=round(px_usd*cur_rate) if px_usd else None
        avg_krw=round(avg_usd*buy_rate) if avg_usd else (round(cost_krw/qty) if qty else None)
        eval_krw=int(val_krw) if val_krw else (round(px_krw*qty) if (px_krw and qty) else None)
        pnl_krw=round(eval_krw-cost_krw) if eval_krw is not None else None
        pct=pnl_krw/cost_krw if (pnl_krw is not None and cost_krw) else None
        result.append({'상품명':nm,'보유수량':qty,'현재가':px_krw if px_krw is not None else '-',
                       '평균매입가':avg_krw if avg_krw is not None else '-','매입금액':cost_krw,
                       '평가금액':eval_krw if eval_krw is not None else '-',
                       '평가손익':pnl_krw if pnl_krw is not None else '-',
                       '손익률':pct if pct is not None else '-'})
    return pd.DataFrame(result, columns=CLEAN_COLS)


def parse_standard_xlsx(path: Path) -> pd.DataFrame:
    try: df = pd.read_excel(path, dtype=str)
    except Exception as e: raise ValueError(f'읽기 실패: {e}')
    df.columns = [str(c).strip() for c in df.columns]
    if '상품명' not in df.columns: raise ValueError("'상품명' 컬럼 없음")
    for col in ['보유수량','현재가','평균매입가','매입금액','평가금액','평가손익','손익률']:
        if col in df.columns: df[col] = df[col].apply(lambda v: _dash(v))
    for col in CLEAN_COLS:
        if col not in df.columns: df[col] = '-'
    df = df[df['상품명'].notna() & (df['상품명'].str.strip() != '')].reset_index(drop=True)
    return df[CLEAN_COLS]


def parse_account_file(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    try:
        if ext in ('.xlsx', '.xls'):
            df = pd.read_excel(path, dtype=str)
        elif ext == '.csv':
            df = pd.read_csv(path, dtype=str, encoding='utf-8-sig')
        else:
            return pd.DataFrame()
    except Exception as e:
        print(f'[WARN] 파일 읽기 실패: {path.name} → {e}')
        return pd.DataFrame()
    df.columns = [str(c).strip() for c in df.columns]
    if '상품명' not in df.columns or '매입금액' not in df.columns:
        return pd.DataFrame()
    for col in ['보유수량','현재가','평균매입가','매입금액','평가금액','평가손익','손익률']:
        if col in df.columns: df[col] = df[col].apply(_to_float)
    df = df[df['상품명'].notna() & (df['상품명'].str.strip() != '')].reset_index(drop=True)
    return df


def parse_filename(name: str) -> dict:
    stem = Path(name).stem
    parts = stem.split('_', 3)
    keys = ['name', 'broker', 'account', 'desc']
    return {k: (parts[i] if i < len(parts) else '') for i, k in enumerate(keys)}


def _detect_broker(stem: str) -> str:
    parts = stem.split('_')
    return parts[1] if len(parts) > 1 else ''

def _is_samsung_format(path: Path) -> bool:
    try:
        from openpyxl import load_workbook as _lw
        ws = _lw(path, data_only=True, read_only=True).active
        for row in ws.iter_rows(max_row=8, values_only=True):
            flat = ' '.join(str(v or '') for v in row)
            if '종목명' in flat and '잔고수량' in flat: return True
    except Exception: pass
    return False


def load_folder(root: str) -> dict:
    DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    root_path = Path(root)
    snapshots: dict = {}

    for d in sorted(root_path.iterdir()):
        if not d.is_dir() or not DATE_RE.match(d.name): continue
        accounts = []
        for f in sorted(d.iterdir()):
            if f.name.startswith('~$'): continue
            if f.name.lower() == 'comments.xlsx': continue
            if f.suffix.lower() not in ('.xlsx', '.xls', '.csv'): continue
            info = parse_filename(f.name)
            df   = parse_account_file(f)
            if df.empty: continue
            info['df']   = df
            info['date'] = d.name
            accounts.append(info)
        if accounts:
            snapshots[d.name] = accounts

    comments: dict = {}
    for d in root_path.iterdir():
        if not d.is_dir() or not DATE_RE.match(d.name): continue
        cp = d / 'comments.xlsx'
        if not cp.exists(): continue
        try:
            cdf  = pd.read_excel(cp, dtype=str, header=None)
            lines = [str(v).strip() for v in cdf.iloc[:, 0]
                     if str(v).strip() and str(v).strip().lower() != 'nan']
            if lines: comments[d.name] = lines
        except Exception as e:
            print(f'[WARN] comments 읽기 실패: {cp.name} → {e}')

    dates = sorted(snapshots)
    return {'dates': dates, 'snapshots': snapshots, 'comments': comments}


def build_summary(data: dict) -> pd.DataFrame:
    rows = []
    for date, accounts in data['snapshots'].items():
        for acc in accounts:
            df    = acc['df']
            cost  = df['매입금액'].sum(skipna=True) if '매입금액' in df.columns else 0.0
            value = df['평가금액'].sum(skipna=True) if '평가금액' in df.columns else 0.0
            rows.append({'date': date, 'name': acc['name'], 'broker': acc['broker'],
                         'account': acc['account'], 'desc': acc['desc'],
                         'cost': cost, 'value': value, 'pnl': value - cost})
    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['date','name','broker','account','desc','cost','value','pnl'])


# ════════════════════════════════════════════════════════════════
# 유틸 함수
# ════════════════════════════════════════════════════════════════

def fmt_won(v, prefix=True) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)): return '-'
    s = f'{abs(v):,.0f}'
    if prefix:
        sign = '+' if v >= 0 else '-'
        return f'{sign}{s}'
    return s

def fmt_pct(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)): return '-'
    return f'{v*100:+.2f}%'

def color_pnl(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)): return '#555'
    return '#C00000' if v < 0 else '#1565C0'

def _pnl_color(pct) -> str:
    if pct is None or (isinstance(pct, float) and pd.isna(pct)): return '#333'
    if pct <= -0.50: return '#7B0000'
    if pct <= -0.30: return '#C62828'
    if pct <= -0.10: return '#E57373'
    if pct <   0.10: return '#333'
    if pct <   0.50: return '#90CAF9'
    if pct <   1.00: return '#1E88E5'
    if pct <   3.00: return '#1565C0'
    return '#0D47A1'


# ── HTML 테이블 생성 헬퍼 ────────────────────────────────────────
_TH_STYLE = ('background:#1E3050;color:white;padding:6px 8px;'
             'font-size:13px;white-space:nowrap;')
_TD_BASE  = 'padding:5px 8px;font-size:13px;white-space:nowrap;border-bottom:1px solid #eee;'

def _html_table(headers, rows, col_aligns=None, row_bg=None, cell_fg=None):
    """
    headers   : list[str]
    rows      : list[list[str]]   (문자열로 변환된 셀 값)
    col_aligns: list['left'|'right'|'center']  (기본 right)
    row_bg    : dict{row_idx: '#색상'}  (행 배경)
    cell_fg   : dict{(row_idx,col_idx): '#색상'}  (셀 글자색)
    """
    col_aligns = col_aligns or ['right'] * len(headers)
    row_bg     = row_bg or {}
    cell_fg    = cell_fg or {}

    th_cells = ''.join(
        f'<th style="{_TH_STYLE}text-align:{al};">{h}</th>'
        for h, al in zip(headers, col_aligns)
    )
    body = ''
    for ri, row in enumerate(rows):
        bg    = row_bg.get(ri, 'white')
        is_bold = ri in row_bg   # 소계/합계 행은 굵게
        td_rows = ''
        for ci, (cell, al) in enumerate(zip(row, col_aligns)):
            fg  = cell_fg.get((ri, ci), '#333')
            bld = 'font-weight:bold;' if is_bold else ''
            td_rows += (f'<td style="{_TD_BASE}text-align:{al};'
                        f'background:{bg};color:{fg};{bld}">{cell}</td>')
        body += f'<tr>{td_rows}</tr>'

    return (f'<div style="overflow-x:auto"><table style="border-collapse:collapse;'
            f'width:100%;font-family:\'Malgun Gothic\',\'NanumGothic\',sans-serif;">'
            f'<thead><tr>{th_cells}</tr></thead><tbody>{body}</tbody></table></div>')


# ════════════════════════════════════════════════════════════════
# GitHub 데이터 로더
# ════════════════════════════════════════════════════════════════

@st.cache_data(ttl=300, show_spinner='GitHub에서 데이터 다운로드 중...')
def load_from_github(token: str, owner: str, repo: str, branch: str = 'main') -> dict:
    """
    GitHub private repo 구조:
      {repo}/
        {YYYY-MM-DD}/
          {이름}_{증권사}_{계좌번호}.xlsx
          comments.xlsx  (선택)
    → 임시 디렉터리에 다운로드 후 load_folder() 호출
    """
    headers_gh = {
        'Authorization': f'token {token}',
        'Accept':        'application/vnd.github.v3+json',
    }
    api_base = f'https://api.github.com/repos/{owner}/{repo}'
    DATE_RE  = re.compile(r'^\d{4}-\d{2}-\d{2}$')

    # 루트 목록
    resp = requests.get(f'{api_base}/contents', headers=headers_gh,
                        params={'ref': branch}, timeout=20)
    resp.raise_for_status()

    tmpdir = tempfile.mkdtemp(prefix='asset_mgr_')

    for item in resp.json():
        if item['type'] != 'dir' or not DATE_RE.match(item['name']):
            continue
        date_dir = Path(tmpdir) / item['name']
        date_dir.mkdir(exist_ok=True)

        # 날짜 폴더 내 파일 목록
        fr = requests.get(item['url'], headers=headers_gh,
                          params={'ref': branch}, timeout=20)
        fr.raise_for_status()

        for fitem in fr.json():
            if fitem['type'] != 'file':
                continue
            ext = Path(fitem['name']).suffix.lower()
            if ext not in ('.xlsx', '.xls', '.csv'):
                continue
            dl = requests.get(fitem['download_url'], headers=headers_gh, timeout=30)
            dl.raise_for_status()
            (date_dir / fitem['name']).write_bytes(dl.content)

    return load_folder(tmpdir)


# ════════════════════════════════════════════════════════════════
# 탭 렌더링
# ════════════════════════════════════════════════════════════════

def render_trend(data: dict, summary: pd.DataFrame):
    if summary is None or summary.empty:
        st.info('데이터가 없습니다.')
        return

    dates   = sorted(summary['date'].unique())
    n_dates = len(dates)
    total   = summary.groupby('date')[['cost','value']].sum().reindex(dates)

    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    fig.patch.set_facecolor('#F0F2F5')

    # ── 전체 자산 추이 ──────────────────────────────────────────
    ax1 = axes[0]; ax1.set_facecolor('white')
    if n_dates == 1:
        items = {'매입원금': total['cost'].iloc[0], '평가금액': total['value'].iloc[0]}
        bars  = ax1.bar(items.keys(), items.values(), color=['#888','#4472C4'],
                        width=0.5, edgecolor='white')
        for b in bars:
            h = b.get_height()
            ax1.text(b.get_x()+b.get_width()/2, h+h*0.01,
                     f'{h/1e8:.2f}억', ha='center', fontsize=10, fontweight='bold')
        ax1.set_title('전체 자산 현황', fontsize=13, fontweight='bold', pad=10)
    else:
        ax1.plot(dates, total['value'], 'o-', color='#4472C4',
                 linewidth=2.5, markersize=8, label='평가금액', zorder=3)
        ax1.plot(dates, total['cost'],  's--', color='#888',
                 linewidth=1.5, markersize=5, label='매입원금', zorder=2)
        ax1.fill_between(dates, total['cost'], total['value'], alpha=0.12, color='#4472C4')
        last_v = total['value'].iloc[-1]; last_c = total['cost'].iloc[-1]
        for d, v, c in zip(dates, total['value'], total['cost']):
            is_last = (d == dates[-1])
            if pd.notna(v):
                ax1.annotate(f'{v/1e8:.2f}억', xy=(d,v), xytext=(0,10),
                             textcoords='offset points', ha='center', fontsize=8,
                             fontweight='bold' if is_last else 'normal', color='#4472C4')
            if pd.notna(c):
                ax1.annotate(f'{c/1e8:.2f}억', xy=(d,c), xytext=(0,-16),
                             textcoords='offset points', ha='center', fontsize=8,
                             fontweight='bold' if is_last else 'normal', color='#888')
        pnl_pct = (last_v-last_c)/last_c*100 if last_c else 0
        ax1.set_title(f'전체 자산 추이  (수익률 {pnl_pct:+.1f}%)',
                      fontsize=13, fontweight='bold', pad=10)
        ax1.tick_params(axis='x', rotation=30); ax1.legend(fontsize=9)
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f'{x/1e8:.1f}억'))
    ax1.grid(axis='y', alpha=0.3); ax1.spines[['top','right']].set_visible(False)

    # ── 인별 추이 ────────────────────────────────────────────────
    ax2 = axes[1]; ax2.set_facecolor('white')
    by_name = summary.groupby(['date','name'])['value'].sum().reset_index()
    names   = sorted(by_name['name'].unique())

    if n_dates == 1:
        vals   = [by_name[by_name['name']==nm]['value'].sum() for nm in names]
        colors = [PERSON_COLORS.get(nm, DEFAULT_COLORS[i]) for i,nm in enumerate(names)]
        bars   = ax2.bar(names, vals, color=colors, width=0.5, edgecolor='white')
        for b in bars:
            h = b.get_height()
            ax2.text(b.get_x()+b.get_width()/2, h+h*0.01,
                     f'{h/1e8:.2f}억', ha='center', fontsize=10, fontweight='bold')
        ax2.set_title('인별 자산 현황', fontsize=13, fontweight='bold', pad=10)
    else:
        for i, nm in enumerate(names):
            sub   = by_name[by_name['name']==nm].set_index('date').reindex(dates)
            color = PERSON_COLORS.get(nm, DEFAULT_COLORS[i])
            ax2.plot(dates, sub['value'], 'o-', color=color,
                     linewidth=2.5, markersize=8, label=nm, zorder=3)
            for d, v in zip(dates, sub['value']):
                if pd.notna(v):
                    ax2.annotate(f'{v/1e8:.2f}억', xy=(d,v), xytext=(0,10),
                                 textcoords='offset points', ha='center', fontsize=8,
                                 fontweight='bold' if d==dates[-1] else 'normal', color=color)
        ax2.set_title('인별 자산 추이', fontsize=13, fontweight='bold', pad=10)
        ax2.tick_params(axis='x', rotation=30); ax2.legend(fontsize=9)
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f'{x/1e8:.1f}억'))
    ax2.grid(axis='y', alpha=0.3); ax2.spines[['top','right']].set_visible(False)

    fig.tight_layout(pad=2.5)
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)


def render_alloc(data: dict, summary: pd.DataFrame, sel: str):
    day = summary[summary['date'] == sel]
    if day.empty:
        st.info('데이터가 없습니다.'); return

    names     = sorted(day['name'].unique())
    n_persons = len(names)
    fig, axes = plt.subplots(2, max(2, n_persons), figsize=(14, 10))
    if axes.ndim == 1: axes = axes.reshape(2, -1)
    fig.patch.set_facecolor('#F0F2F5')
    fig.suptitle(f'{sel} 기준 자산 구성', fontsize=13, fontweight='bold')

    def stock_series(filter_name=None, top_n=15):
        vals = {}
        for acc in data['snapshots'].get(sel, []):
            if filter_name and acc['name'] != filter_name: continue
            for _, row in acc['df'].iterrows():
                nm  = str(row.get('상품명','')).strip()
                val = row.get('평가금액')
                if nm and isinstance(val,(int,float)) and not pd.isna(val) and val>0:
                    vals[nm] = vals.get(nm,0) + val
        s = pd.Series(vals).sort_values(ascending=False)
        if len(s) > top_n:
            s = pd.concat([s.iloc[:top_n], pd.Series({'기타': s.iloc[top_n:].sum()})])
        return s

    def draw_pie(ax, series, title):
        vals = series[series > 0].sort_values(ascending=False)
        if vals.empty:
            ax.text(0.5,0.5,'데이터 없음',ha='center',va='center',transform=ax.transAxes)
            ax.set_title(title, fontsize=11, fontweight='bold', pad=6); return
        colors = [PERSON_COLORS.get(idx, DEFAULT_COLORS[i % len(DEFAULT_COLORS)])
                  for i,idx in enumerate(vals.index)]
        _, texts, auts = ax.pie(vals.values, labels=vals.index, autopct='%1.1f%%',
                                startangle=90, colors=colors, pctdistance=0.78,
                                wedgeprops={'linewidth':1.5,'edgecolor':'white'})
        for t in texts: t.set_fontsize(8)
        for t in auts:  t.set_fontsize(7.5)
        ax.text(0,0,f'{vals.sum()/1e8:.2f}억',ha='center',va='center',
                fontsize=10,fontweight='bold',color='#333')
        ax.set_title(title, fontsize=11, fontweight='bold', pad=6)

    draw_pie(axes[0,0], day.groupby('name')['value'].sum(), '인별 구성')
    draw_pie(axes[0,1], stock_series(filter_name=None), '전체 종목별 구성')
    for col in range(2, axes.shape[1]): axes[0,col].set_visible(False)

    for col, nm in enumerate(names):
        color = PERSON_COLORS.get(nm, DEFAULT_COLORS[col % len(DEFAULT_COLORS)])
        draw_pie(axes[1,col], stock_series(filter_name=nm), f'{nm} 종목별 구성')
        axes[1,col].set_title(f'{nm} 종목별 구성', fontsize=11,
                               fontweight='bold', color=color, pad=6)
    for col in range(n_persons, axes.shape[1]): axes[1,col].set_visible(False)

    fig.tight_layout(pad=2.5)
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)


def render_account(data: dict, summary: pd.DataFrame, sel: str):
    day = summary[summary['date'] == sel].copy()
    if day.empty:
        st.info('데이터가 없습니다.'); return

    total_cost  = day['cost'].sum()
    total_value = day['value'].sum()
    total_pnl   = day['pnl'].sum()
    total_pct   = total_pnl / total_cost if total_cost else 0

    # ── 요약 카드 ──────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    def metric_card(col, label, value, color):
        col.markdown(
            f'<div style="background:white;border-radius:8px;padding:14px;text-align:center;">'
            f'<div style="font-size:12px;color:#777;">{label}</div>'
            f'<div style="font-size:18px;font-weight:bold;color:{color};">{value}</div></div>',
            unsafe_allow_html=True)
    metric_card(c1, '총 평가금액',  f'{total_value:,.0f} 원', '#1A4080')
    metric_card(c2, '총 매입원금',  f'{total_cost:,.0f} 원',  '#555')
    metric_card(c3, '평가 손익',    f'{total_pnl:+,.0f} 원',  color_pnl(total_pnl))
    metric_card(c4, '전체 수익률',  f'{total_pct*100:+.2f}%', color_pnl(total_pct))
    st.write('')

    # ── 계좌 테이블 ────────────────────────────────────────────
    headers   = ['이름','증권사','계좌번호','설명','매입원금','평가금액','평가손익','수익률']
    col_aligns= ['center','center','center','left','right','right','right','right']
    pnl_col   = headers.index('평가손익')
    pct_col   = headers.index('수익률')

    day_sorted = day.sort_values(['name','value'], ascending=[True, False])
    rows, row_bg, cell_fg = [], {}, {}
    ri = 0

    for nm in day_sorted['name'].unique():
        person_rows  = day_sorted[day_sorted['name'] == nm]
        person_cost  = person_rows['cost'].sum()
        person_value = person_rows['value'].sum()
        person_pnl   = person_rows['pnl'].sum()
        person_pct   = person_pnl / person_cost if person_cost else 0

        for _, r in person_rows.iterrows():
            pnl = r['pnl']; pct = pnl / r['cost'] if r['cost'] else 0
            rows.append([r['name'], r['broker'], r['account'], r['desc'],
                         f"{r['cost']:,.0f}", f"{r['value']:,.0f}",
                         f"{pnl:+,.0f}", f"{pct*100:+.2f}%"])
            cell_fg[(ri, pnl_col)] = _pnl_color(pct)
            cell_fg[(ri, pct_col)] = _pnl_color(pct)
            ri += 1

        # 인별 소계
        rows.append([f'[{nm} 소계]', '', '', '',
                     f'{person_cost:,.0f}', f'{person_value:,.0f}',
                     f'{person_pnl:+,.0f}', f'{person_pct*100:+.2f}%'])
        row_bg[ri] = '#DCE6F1'
        cell_fg[(ri, pnl_col)] = _pnl_color(person_pct)
        cell_fg[(ri, pct_col)] = _pnl_color(person_pct)
        ri += 1

    # 합계
    rows.append(['【합계】', '', '', '',
                 f'{total_cost:,.0f}', f'{total_value:,.0f}',
                 f'{total_pnl:+,.0f}', f'{total_pct*100:+.2f}%'])
    row_bg[ri] = '#D8EAD3'
    cell_fg[(ri, pnl_col)] = _pnl_color(total_pct)
    cell_fg[(ri, pct_col)] = _pnl_color(total_pct)

    st.markdown(_html_table(headers, rows, col_aligns, row_bg, cell_fg),
                unsafe_allow_html=True)

    # ── 메모 (comments + 수량 변화) ─────────────────────────────
    comments   = data.get('comments', {}).get(sel, [])
    qty_changes= []
    dates_list = data['dates']
    sel_idx    = dates_list.index(sel) if sel in dates_list else -1

    if sel_idx > 0:
        prev_snap  = data['snapshots'].get(dates_list[sel_idx - 1], [])
        curr_snap  = data['snapshots'].get(sel, [])
        prev_map, curr_map, price_map = {}, {}, {}
        for acc in prev_snap:
            for _, row in acc['df'].iterrows():
                qty = row.get('보유수량')
                if isinstance(qty,(int,float)) and not pd.isna(qty):
                    key = (acc['name'],acc['broker'],acc['account'],str(row.get('상품명','')))
                    prev_map[key] = qty
        for acc in curr_snap:
            for _, row in acc['df'].iterrows():
                qty = row.get('보유수량')
                if isinstance(qty,(int,float)) and not pd.isna(qty):
                    key = (acc['name'],acc['broker'],acc['account'],str(row.get('상품명','')))
                    curr_map[key] = qty
                    price = row.get('현재가')
                    if isinstance(price,(int,float)) and not pd.isna(price):
                        price_map[key] = price
        def _fq(q): return f'{q:,.0f}' if q%1==0 else f'{q:,.2f}'
        for key in sorted(set(prev_map)|set(curr_map), key=lambda k: k[3]):
            p, c = prev_map.get(key,0), curr_map.get(key,0)
            if p == c: continue
            name, broker, account, stock = key
            direction = '매수' if c > p else '매도'
            delta     = abs(c - p)
            curr_price= price_map.get(key)
            amount    = delta * curr_price if curr_price is not None else 0
            detail    = f'{_fq(p)}주 --> {_fq(c)}주'
            if curr_price is not None: detail += f':    {amount:,.0f}원'
            text = f'[{name}][{broker}][{account}] {stock} {direction} ({detail})'
            qty_changes.append((text, amount))

    if comments or qty_changes:
        st.write('')
        memo_lines = []
        for c in comments:
            memo_lines.append(f'<span style="color:#333;">• {c}</span>')
        if comments and qty_changes:
            memo_lines.append('')
        for text, amount in qty_changes:
            if amount >= 30_000_000:
                color = '#C00000'; weight = 'bold'
            elif amount >= 10_000_000:
                color = '#D46000'; weight = 'bold'
            else:
                color = '#333'; weight = 'normal'
            memo_lines.append(
                f'<span style="color:{color};font-weight:{weight};">• {text}</span>')
        memo_html = '<br>'.join(memo_lines)
        st.markdown(
            f'<div style="background:#FFFDE7;border-radius:8px;padding:12px 16px;'
            f'font-family:\'Malgun Gothic\',\'NanumGothic\',sans-serif;font-size:13px;'
            f'line-height:1.8;">'
            f'<b style="color:#7A6500;">📝 메모</b><br>{memo_html}</div>',
            unsafe_allow_html=True)


def render_stock(data: dict, sel: str):
    accounts = data['snapshots'].get(sel, [])
    if not accounts:
        st.info('데이터가 없습니다.'); return

    dates   = data['dates']
    sel_idx = dates.index(sel) if sel in dates else -1
    prev_qty: dict = {}
    if sel_idx > 0:
        for acc in data['snapshots'].get(dates[sel_idx-1], []):
            for _, row in acc['df'].iterrows():
                qty = row.get('보유수량')
                if qty is not None and not (isinstance(qty,float) and pd.isna(qty)):
                    prev_qty[(acc['name'],acc['account'],str(row.get('상품명','')))] = qty

    rows_data = []
    for acc in accounts:
        for _, r in acc['df'].iterrows():
            rows_data.append({'이름':acc['name'],'증권사':acc['broker'],'계좌':acc['account'],
                              '종목명':r.get('상품명',''),'수량':r.get('보유수량'),
                              '현재가':r.get('현재가'),'평균매입가':r.get('평균매입가'),
                              '매입금액':r.get('매입금액'),'평가금액':r.get('평가금액'),
                              '평가손익':r.get('평가손익'),'손익률':r.get('손익률')})
    all_df = pd.DataFrame(rows_data)

    # ── 수익률 TOP/BOTTOM 차트 ──────────────────────────────────
    valid = all_df.dropna(subset=['손익률','평가금액']).copy()
    valid = valid[valid['평가금액'] > 0]
    if not valid.empty:
        top5     = valid.nlargest(5,'손익률')
        bot5     = valid.nsmallest(5,'손익률')
        chart_df = pd.concat([top5,bot5]).drop_duplicates(subset=['이름','계좌','종목명'])
        chart_df = chart_df.sort_values('손익률')
        chart_df['label'] = chart_df['종목명']

        fig, ax = plt.subplots(figsize=(13, 3.8))
        fig.patch.set_facecolor('#F0F2F5')
        ax.set_facecolor('white')
        colors = ['#C00000' if v<0 else '#1D6B2B' for v in chart_df['손익률']]
        bars   = ax.barh(chart_df['label'], chart_df['손익률']*100,
                         color=colors, edgecolor='white', height=0.65)
        for bar, val in zip(bars, chart_df['손익률']*100):
            x = bar.get_width()
            ax.text(x+(1 if x>=0 else -1), bar.get_y()+bar.get_height()/2,
                    f'{val:+.1f}%', va='center',
                    ha='left' if x>=0 else 'right', fontsize=8)
        ax.axvline(0, color='#555', linewidth=0.8)
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f'{x:.0f}%'))
        ax.set_title('수익률 TOP 5 / BOTTOM 5', fontsize=11, fontweight='bold')
        ax.grid(axis='x', alpha=0.25); ax.spines[['top','right']].set_visible(False)
        fig.tight_layout(pad=1.5)
        st.pyplot(fig, use_container_width=True)
        plt.close(fig)

    # ── 종목 테이블 ────────────────────────────────────────────
    headers    = ['이름','증권사','계좌','종목명','수량','현재가','평균매입가',
                  '매입금액','평가금액','평가손익','수익률']
    col_aligns = ['center','center','center','left','right','right','right',
                  'right','right','right','right']
    pnl_col = headers.index('평가손익')
    pct_col = headers.index('수익률')
    qty_col = headers.index('수량')

    sort_df = all_df.sort_values(['이름','평가금액'], ascending=[True,False], na_position='last')
    tbl_rows, cell_fg = [], {}

    for ri, (_, r) in enumerate(sort_df.iterrows()):
        pnl = r['평가손익']; pct = r['손익률']
        qty_str = f"{r['수량']:,.0f}" if pd.notna(r['수량']) else '-'
        row_vals = [
            r['이름'], r['증권사'], r['계좌'], r['종목명'],
            qty_str,
            f"{r['현재가']:,.0f}"     if pd.notna(r['현재가'])     else '-',
            f"{r['평균매입가']:,.0f}" if pd.notna(r['평균매입가']) else '-',
            f"{r['매입금액']:,.0f}"   if pd.notna(r['매입금액'])   else '-',
            f"{r['평가금액']:,.0f}"   if pd.notna(r['평가금액'])   else '-',
            f"{pnl:+,.0f}"            if pd.notna(pnl)             else '-',
            f"{pct*100:+.2f}%"        if pd.notna(pct)             else '-',
        ]
        tbl_rows.append(row_vals)
        fg = _pnl_color(pct) if pd.notna(pct) else color_pnl(pnl)
        if pd.notna(pnl): cell_fg[(ri, pnl_col)] = fg
        if pd.notna(pct): cell_fg[(ri, pct_col)] = fg
        # 수량 변화 → 초록색
        if prev_qty and qty_str != '-':
            key = (r['이름'], r['계좌'], str(r['종목명']))
            if key not in prev_qty or prev_qty[key] != r['수량']:
                cell_fg[(ri, qty_col)] = '#2E7D32'

    st.markdown(_html_table(headers, tbl_rows, col_aligns, cell_fg=cell_fg),
                unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# 인증 설정 (st.secrets 기반)
# ════════════════════════════════════════════════════════════════

def _build_credentials() -> dict:
    """st.secrets 에서 credentials 딕셔너리 구성."""
    creds = {'usernames': {}}
    for uname, udata in st.secrets.get('credentials', {}).get('usernames', {}).items():
        creds['usernames'][uname] = {
            'name':     udata.get('name', uname),
            'password': udata.get('password', ''),
        }
    return creds


# ════════════════════════════════════════════════════════════════
# 메인
# ════════════════════════════════════════════════════════════════

def main():
    st.set_page_config(
        page_title='개인 자산관리',
        page_icon='📈',
        layout='wide',
        initial_sidebar_state='expanded',
    )

    # ── 전역 CSS ─────────────────────────────────────────────────
    st.markdown("""
    <style>
    [data-testid="stSidebar"] { background: #1E3050; }
    [data-testid="stSidebar"] * { color: #CCC !important; }
    .stTabs [data-baseweb="tab-list"] { gap: 4px; }
    .stTabs [data-baseweb="tab"] { padding: 8px 20px; font-size: 15px; }
    div[data-testid="metric-container"] { background: white; border-radius: 8px; padding: 10px; }
    </style>""", unsafe_allow_html=True)

    # ── 인증 ─────────────────────────────────────────────────────
    try:
        credentials  = _build_credentials()
        cookie_cfg   = st.secrets.get('cookie', {})
        authenticator = stauth.Authenticate(
            credentials,
            cookie_cfg.get('name',        'asset_mgr'),
            cookie_cfg.get('key',         'default_secret_key'),
            int(cookie_cfg.get('expiry_days', 30)),
        )
    except Exception as e:
        st.error(f'인증 설정 오류: {e}')
        st.stop()

    authenticator.login(location='main')

    auth_status = st.session_state.get('authentication_status')
    name        = st.session_state.get('name', '')

    if auth_status is False:
        st.error('아이디 또는 비밀번호가 올바르지 않습니다.')
        st.stop()
    if auth_status is None:
        st.warning('아이디와 비밀번호를 입력해 주세요.')
        st.stop()

    # ── 로그인 성공 ───────────────────────────────────────────────
    with st.sidebar:
        st.markdown(f'## 📈 자산관리')
        st.markdown(f'**{name}** 님 환영합니다')
        authenticator.logout(location='sidebar')
        st.divider()

        # 데이터 로드
        if st.button('🔄  데이터 새로고침', use_container_width=True):
            st.cache_data.clear()
            st.rerun()

    # ── GitHub 에서 데이터 로드 ───────────────────────────────────
    try:
        gh = st.secrets['github']
        data = load_from_github(
            token  = gh['token'],
            owner  = gh['owner'],
            repo   = gh['repo'],
            branch = gh.get('branch', 'main'),
        )
    except Exception as e:
        st.error(f'데이터 로드 실패: {e}')
        st.stop()

    if not data['dates']:
        st.warning('날짜 폴더(YYYY-MM-DD 형식)가 없습니다.')
        st.stop()

    summary = build_summary(data)

    # ── 사이드바: 날짜 선택 ───────────────────────────────────────
    with st.sidebar:
        sel_date = st.selectbox(
            '기준일 선택',
            options=data['dates'],
            index=len(data['dates']) - 1,
        )
        n_dates = len(data['dates'])
        n_files = sum(len(v) for v in data['snapshots'].values())
        st.caption(f'날짜: {n_dates}개 | 파일: {n_files}개')
        st.caption(f'{data["dates"][0]} ~ {data["dates"][-1]}')

    # ── 탭 ────────────────────────────────────────────────────────
    t1, t2, t3, t4 = st.tabs(['📈  자산 추이', '🥧  자산 구성',
                                '🏦  계좌 현황', '📊  종목 현황'])
    with t1:
        render_trend(data, summary)
    with t2:
        render_alloc(data, summary, sel_date)
    with t3:
        render_account(data, summary, sel_date)
    with t4:
        render_stock(data, sel_date)


if __name__ == '__main__':
    main()


